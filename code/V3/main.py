import sys

from openpyxl.styles import Alignment

import GUI
from PyQt5.QtWidgets import QApplication, QMainWindow
from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
import os
import time
from openpyxl import Workbook, load_workbook
from tkinter import messagebox, Tk, simpledialog, Message
import random
from threading import Lock, Thread
import requests
import login_window as lw


if not os.path.exists('icon.png'):
    r = requests.get('https://aero80wd.github.io/icon.png')
    with open('icon.png', "wb") as f:
        for chunk in r.iter_content(chunk_size=512):
            f.write(chunk)

nums = None
ws = None
entry = 0
from PyQt5 import QtCore

QtCore.QCoreApplication.setAttribute(QtCore.Qt.AA_EnableHighDpiScaling)


def sheet_model(sheet, files_name=False):
    if not files_name:

        for x in os.listdir('./Data'):
            file = load_workbook('./Data/' + x)
            hh = file.create_sheet(sheet)
            hh.merge_cells('A1:C1')
            hh['A1'] = '双十中学' + sheet
            hh['A2'] = '号数'
            hh['B2'] = '成绩'
            hh['C2'] = '等级'
            hh['A1'].alignment = Alignment(horizontal='center', vertical='center')
            file.save('./Data/' + x)
        return
    if files_name:
        file = load_workbook('.\\Data\\' + files_name + '.xlsx')
        hh = file[str(sheet)]
        hh.merge_cells('A1:C1')
        hh['A1'] = '双十中学' + sheet
        hh['A2'] = '号数'
        hh['B2'] = '成绩'
        hh['C2'] = '等级'
        hh['A1'].alignment = Alignment(horizontal='center', vertical='center')
        file.save('.\\Data\\' + files_name + '.xlsx')
        return True


if not os.path.exists('Data'):
    os.mkdir('./Data')

if __name__ == "__main__":
    app = QApplication(sys.argv)

    file = None

    MainWindow = QMainWindow()
    ui = GUI.Ui_MainWindow()
    extra = {

        # Button colors
        'danger': '#dc3545',
        'warning': '#ffc107',
        'success': '#7DC2B0',

        # Font
        'font_family': 'OPPOSans L', 'font_size': '5px',
        'line_height': '1px',
    }
    MainWindow.setWindowOpacity(0.3)
    ui.setupUi(MainWindow)
    MainWindow.setWindowTitle('BANJ-SOUCE V3 (Build FRI 113)')
    MainWindow.setWindowIcon(QIcon('icon.png'))
    file_list = os.listdir('Data')

    ui.banje.setPlaceholderText('请输入考试名')

def reset_pwd():
    inputnopwd = simpledialog.askstring('reset pwssword','input old password:')
    with open('password.txt','r') as npf:
        if npf.readline().strip('\'[]') != inputnopwd:
            messagebox.showerror('error','password error!')
            return
    inputnewpwd = simpledialog.askstring('reset password','input new password')
    inputnewpwd2 = simpledialog.askstring('reset password', 'input new password again')
    if inputnewpwd == inputnewpwd2:
        with open('password.txt', 'w') as npf:
            pass
        with open('password.txt', 'w') as npf:
            npf.write(inputnewpwd)
    if inputnewpwd != inputnewpwd2:
        messagebox.showerror('error','input 2 password inconsistency!')
        return
    messagebox.showinfo('info','password reset finish!')
    return
def hidden():
    if not os.path.exists('password.txt'):
        with open('password.txt','w') as newp:
            pass
    with open('password.txt','r') as fpw:
        fpwinf = fpw.readline()
    if fpwinf == '':
        while True:
            inputpwd = simpledialog.askstring('login', 'SetPassword(not input"[]"):')
            if inputpwd != '' and inputpwd != None:
                with open('password.txt', 'w') as fpw:
                    fpw.write(inputpwd)
                    break
    ui.new_file_entry.clear()
    ui.new_sheet_3.clear()
    ui.open_file_entry.clear()
    ui.tableWidget.setRowCount(0)
    ui.str_is.clear()
    ui.banje.clear()
    ui.lineEdit.clear()
    lw.path = 0
    MainWindow.setHidden(True)
    lw.uishow()
    lw.Dialog.exec_()
    if lw.path == 0:
        exit()
    if lw.path == 1:
        MainWindow.setHidden(False)


hidden()

for x in file_list:
    if '.xlsx' in x == False:
        file_list.remove(x)
    if '.xlsx' in x:
        ui.file.addItem(x)


def files_view():
    while True:
        global file_list

        ui.file.clear()
        file_list = os.listdir('Data')
        for x in file_list:

            if '.xlsx' not in x:
                file_list.remove(x)
            if '.xlsx' in x:
                ui.file.addItem(x)
        ui.label.setText(str(len(file_list)) + ' 班级')
        time.sleep(0.3)


t1 = Thread(target=files_view)
t1.start()


def open_data_files_name():
    os.system('start .\\Data')


def new_file():
    if ui.new_file_entry.text() == '':
        messagebox.showerror('错误', '创建失败')
        return

    name = ui.new_file_entry.text()
    temp = Workbook()
    temp.save('.\\Data\\' + str(name) + '.xlsx')
    sheet_model('Sheet', ui.new_file_entry.text())
    ui.new_file_entry.clear()
    messagebox.showinfo('提示', '创建成功')


def caidan_2():
    r = range(random.randint(100, 10000))
    print(list(r)[-1])
    for x in r:
        messagebox.showinfo('CD' + str(x), 'abcdefghijklmnopqrstuvwxyz'[random.randint(0, 25)])


def dali_cai():
    rtext = ['bilibili', 'banj-souce', 'build', 'video', 'python', 'python-qt', 'pycharm', 'windows']
    print(ui.dial.value())
    if (ui.dial.value() >= 90) and (ui.dial.value() < 99):
        ui.labelC.setText(random.choice(rtext))
    elif ui.dial.value() == 99:
        caidan_2()
    else:
        ui.labelC.setText('')


def open_file(text=False, bool=True):
    global nums
    global ws
    global file
    global entry
    try:
        if text == False:
            entry = ui.open_file_entry.text().split(';')
        else:
            entry = text.split(';')
        if bool == True:
            file = load_workbook('.\\Data\\' + str(entry[0]) + '.xlsx')
            ui.open_text.setText('.\\Data\\' + str(entry[0]) + '.xlsx;' + entry[1])
        if bool == False:
            file = load_workbook('.\\Data\\' + str(entry[0]))

        ws = file[str(entry[1])]
        if bool == True:

            ui.tableWidget.setRowCount(ws.max_row - 2)
            for x in range(ws.max_row - 2):
                a = ws['A' + str(x + 3)].value
                b = ws['B' + str(x + 3)].value
                c = ws['C' + str(x + 3)].value
                ui.tableWidget.setItem(x, 0, QTableWidgetItem(str(a)))
                ui.tableWidget.setItem(x, 1, QTableWidgetItem(str(b)))
                ui.tableWidget.setItem(x, 2, QTableWidgetItem(str(c)))
        nums = 0

        for x in range(ws.max_row - 2):
            nums += ws['B' + str(x + 3)].value

        print(0)
        return nums
    except:
        print(-1)
        if bool:
            messagebox.showerror('错误', '打开失败，有可能是没有输sheet')
        return


def all_nums():
    if str(nums) != 'None':
        messagebox.showinfo('提示', '总分为：' + str(nums))
    else:
        messagebox.showerror('错误', '未打开文件')


def num_1():
    try:
        messagebox.showinfo('提示', '平均分为：' + str(nums / (ws.max_row - 2)))
    except:

        messagebox.showerror('错误', '未打开文件')


def new_sheet():
    try:
        entry = ui.new_sheet_3.text()
        if not entry:
            messagebox.showerror('错误', '创建失败')
            return

        sheet_model(entry)
        messagebox.showinfo('提示', '创建成功')
    except:
        messagebox.showerror('错误', '创建失败')
        return


def num_edit():
    global ws, file
    try:
        nums_2 = simpledialog.askstring('info', 'entry:')
        if nums_2 == None or nums_2 == '':
            return
        if ';' in nums_2:
            file = load_workbook(ui.open_text.text().split(';')[0])
            ws = file[ui.open_text.text().split(';')[1]]
            print(ui.open_text.text().split(';'))
            nums_2 = nums_2.split(';')
            ws['A' + str(int(nums_2[0]) + 2)].value = int(nums_2[0])
            ws['B' + str(int(nums_2[0]) + 2)].value = float(nums_2[1])
            if float(nums_2[1]) >= 90:
                ws['C' + str(int(nums_2[0]) + 2)].value = 'A'
            if 80 <= float(nums_2[1]) < 90:
                ws['C' + str(int(nums_2[0]) + 2)].value = 'B'
            if 60 <= float(nums_2[1]) < 80:
                ws['C' + str(int(nums_2[0]) + 2)].value = 'C'
            if float(nums_2[1]) < 60:
                ws['C' + str(int(nums_2[0]) + 2)].value = 'D'
            if float(nums_2[1]) > 100:
                ws['C' + str(int(nums_2[0]) + 2)].value = '?'
                ws['B' + str(int(nums_2[0]) + 2)].value = 'Max100'
            if float(nums_2[1]) == 100:
                ws['C' + str(int(nums_2[0]) + 2)].value = 'A+'
            print(ws['C' + str(int(nums_2[0]) + 2)])

            file.save(ui.open_text.text().split(';')[0])
            ui.tableWidget.setRowCount(ws.max_row - 2)
            for x in range(ws.max_row - 2):
                a = ws['A' + str(x + 3)].value
                b = ws['B' + str(x + 3)].value
                c = ws['C' + str(x + 3)].value
                ui.tableWidget.setItem(x, 0, QTableWidgetItem(str(a)))
                ui.tableWidget.setItem(x, 1, QTableWidgetItem(str(b)))
                ui.tableWidget.setItem(x, 2, QTableWidgetItem(str(c)))
            nums = 0
            for x in range(ws.max_row - 2):
                nums += ws['B' + str(x + 3)].value
            return

        if ';' not in nums_2:
            ws['A' + str(ws.max_row + 1)].value = ws.max_row -1
            ws['B' + str(ws.max_row)].value = float(nums_2)
            if float(nums_2) >= 90:
                ws['C' + str(ws.max_row)].value = 'A'
            if 80 <= float(nums_2) < 90:
                ws['C' + str(ws.max_row)].value = 'B'
            if 60 <= float(nums_2) < 80:
                ws['C' + str(ws.max_row)].value = 'C'
            if float(nums_2) < 60:
                ws['C' + str(ws.max_row)].value = 'D'
            if float(nums_2) > 100:
                ws['C' + str(ws.max_row)].value = '?'
                ws['B' + str(ws.max_row)].value = 'Max100'
            if float(nums_2) == 100:
                ws['C' + str(ws.max_row)].value = 'A+'

            file.save(ui.open_text.text().split(';')[0])
            ui.tableWidget.setRowCount(ws.max_row-2)
            for x in range(ws.max_row - 2):
                a = ws['A' + str(x + 3)].value
                b = ws['B' + str(x + 3)].value
                c = ws['C' + str(x + 3)].value
                ui.tableWidget.setItem(x, 0, QTableWidgetItem(str(a)))
                ui.tableWidget.setItem(x, 1, QTableWidgetItem(str(b)))
                ui.tableWidget.setItem(x, 2, QTableWidgetItem(str(c)))
            nums = 0
            for x in range(ws.max_row - 2):
                nums += ws['B' + str(x + 3)].value
    except Exception as exp:
        messagebox.showerror('错误', '创建失败' + str(exp))
        return


def str_start():
    try:
        num = None
        entry = ui.str_is.text().split(';')
        if entry[0] == '<':
            num = {}
            for x in range(ws.max_row - 2):
                if int(ws['B' + str(x + 3)].value) < int(entry[1]):
                    num[ws['A' + str(x + 3)].value] = ws['B' + str(x + 3)].value
        if entry[0] == '>':
            num = {}
            for x in range(ws.max_row - 2):
                if int(ws['B' + str(x + 3)].value) >= int(entry[1]):
                    num[ws['A' + str(x + 3)].value] = ws['B' + str(x + 3)].value
        messagebox.showinfo('提示', '号数:' + str(list(num.keys())) + '\n成绩:' + str(list(num.values())))
    except:
        messagebox.showerror('错误', '创建失败')
        return


def banj_f():
    try:
        pj_nums = {}
        for x in os.listdir('./Data'):
            print(open_file(x + ';' + ui.banje.text(), False))
            pj_nums[int(x[:-5])] = open_file(x + ';' + ui.banje.text(), False) / (ws.max_row - 2)
        pj_nums = dict(sorted(pj_nums.items(), key=lambda d: d[1], reverse=True))
        messagebox.showinfo('提示', '班级:' + str(list(pj_nums.keys())) + '\n平均分:' + str(list(pj_nums.values())))
    except Exception as e:

        messagebox.showerror('提示', '错误:' + str(e) + ',应该是输错考试名了')
        return


def banj_pm():
    try:
        banj_p = {}
        for x in range(ws.max_row - 2):
            banj_p[x + 1] = ws['B' + str(x + 3)].value
        banj_p = dict(sorted(banj_p.items(), key=lambda d: d[1], reverse=True))
        messagebox.showinfo('提示', '号数:' + str(list(banj_p.keys())) + '\n分数:' + str(list(banj_p.values())))
    except Exception as e:

        messagebox.showerror('提示', '错误:' + str(e) + ',应该是没打开考试')


def chk_source():
    try:
        chk = {}
        for x in file.sheetnames:
            ttt = file[x]
            chk[x] = ttt['B' + str(int(ui.lineEdit.text()) + 2)].value
        messagebox.showinfo('提示', '考试:' + str(list(chk.keys())) + '\n分数:' + str(list(chk.values())))
    except:
        messagebox.showerror('错误', '查询失败')


ui.pushButton_2.clicked.connect(open_data_files_name)
ui.new_file_2.clicked.connect(new_file)
ui.open_file.clicked.connect(open_file)
ui.all_num.clicked.connect(all_nums)
ui.nums_1.clicked.connect(num_1)
ui.new_sheet_2.clicked.connect(new_sheet)
ui.num_edit.clicked.connect(num_edit)
ui.str_start.clicked.connect(str_start)
ui.banj.clicked.connect(banj_f)
ui.pushButton_3.clicked.connect(banj_pm)
ui.pushButton_4.clicked.connect(chk_source)
ui.pushButton_5.clicked.connect(dali_cai)
ui.logoff.clicked.connect(hidden)
ui.resetpwd.clicked.connect(reset_pwd)
MainWindow.setFixedSize(MainWindow.width(), MainWindow.height())

MainWindow.show()

app.exec_()
MainWindow.close()
